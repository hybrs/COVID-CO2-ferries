"""
    This software is licensed under a BSD license
    Copyright (c) 2022, CMCC

    Redistribution and use in source and binary forms, with or without modification,
     are permitted provided that the following conditions are met:

    Redistributions of source code must retain the above copyright notice,
    this list of conditions and the following disclaimer.
    
    Redistributions in binary form must reproduce the above copyright notice,
    this list of conditions and the following disclaimer in the documentation and/or
    other materials provided with the distribution.
    
    THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED
    WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR
    PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT OWNER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT,
    INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF
    SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY
    THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING
    IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE

"""


import copy
import random
import string
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import pandas as pd
from os import makedirs, path
import datetime
from matplotlib.ticker import MultipleLocator, FuncFormatter
# Excel libraries
from openpyxl import load_workbook
from openpyxl.styles import Color, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
# MLM library and scores
import statsmodels.formula.api as smf
from pymer4.models import Lmer
from sklearn.model_selection import KFold
from sklearn.metrics import mean_squared_error, r2_score
import rpy2.robjects as robjects


# Size of elements in plots
incr = 8
title_fsize = 17+incr
ticks_fsize = 15+incr
label_fsize = 16+incr
txt_fsize   = 13+incr

figsize = (12, 9)


def get_SU_models(dep_var):
    """
    Returns a list of strings, each of which is a formula of a model in [1]. 
    The function takes one argument, which is the name of the dependent variable. 

    Parameters
    ----------
    dep_var : str
        name of the dependent variable, can be `'Etot'` or `'Eber'`

    Returns
    -------
        : list of str
        A list of strings, each string is a model formula.

    References
    ----------

    [1] Mannarini G, Salinas ML, Carelli L, Fassò A. 
    How COVID-19 Affected GHG Emissions of Ferries in Europe. 
    Sustainability. 2022; 14(9):5287. https://doi.org/10.3390/su14095287 


    """

    return [
    '%s ~ 1' % dep_var,
    '%s ~ Dom' % dep_var,
    '%s ~ COVID' % dep_var,
    '%s ~ nCalls' % dep_var,
    '%s ~ COVID*nCalls + Dom' % dep_var,
    '%s ~ VType' % dep_var,
    '%s ~ COVID*VType' % dep_var,
    '%s ~ COVID*VType + COVID*Dom' % dep_var,
    '%s ~ COVID*nCalls + COVID*VType' % dep_var,
    '%s ~ COVID*VType + COVID*Dom + COVID*nCalls' % dep_var,  # old #25 chosen model linear version ********
    '%s ~ COVID*nCalls + COVID*VType + nCalls*VType' % dep_var,
    '%s ~ COVID*nCalls + VType' % dep_var,
    '%s ~ COVID + COVID:nCalls + nCalls:VType + VType' % dep_var,
    '%s ~ COVID + COVID:nCalls + COVID:VType + VType' % dep_var,
    '%s ~ Dom + COVID + nCalls + VType' % dep_var,
    '%s ~ Dom + COVID*nCalls + VType' % dep_var,
    '%s ~ Dom + COVID + nCalls*VType' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + VType' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + nCalls*VType' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + COVID*VType' % dep_var,
    '%s ~ 1 + (1|IMOn)' % dep_var,
    '%s ~ Dom + (1|IMOn)' % dep_var,
    '%s ~ COVID + (1|IMOn)' % dep_var,
    '%s ~ nCalls + (1|IMOn)' % dep_var,
    '%s ~ COVID*nCalls + Dom + (1|IMOn)' % dep_var,
    '%s ~ VType + (1|IMOn)' % dep_var,
    '%s ~ COVID*VType + (1|IMOn)' % dep_var,
    '%s ~ COVID*VType +  COVID*Dom + (1 | IMOn)' % dep_var,
    '%s ~ COVID*nCalls + COVID*VType + (1|IMOn)' % dep_var,
    '%s ~ COVID*VType + COVID*Dom + COVID*nCalls + (1|IMOn)' % dep_var, # old #25 chosen model ********
    '%s ~ COVID*nCalls + COVID*VType + nCalls*VType + (1|IMOn)' % dep_var,
    '%s ~ COVID*nCalls + VType + (1|IMOn)' % dep_var,
    '%s ~ COVID + COVID:nCalls + nCalls:VType + VType + (1|IMOn)' % dep_var,
    '%s ~ COVID + COVID:nCalls + COVID:VType + VType + (1|IMOn)' % dep_var,
    '%s ~ Dom + COVID + nCalls + VType + (1|IMOn)' % dep_var,
    '%s ~ Dom + COVID*nCalls + VType + (1|IMOn)' % dep_var,
    '%s ~ Dom + COVID + nCalls*VType + (1|IMOn)' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + VType + (1|IMOn)' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + nCalls*VType  + (1|IMOn)' % dep_var,
    '%s ~ Dom*COVID*nCalls - Dom:COVID:nCalls + COVID*VType + (1|IMOn)' % dep_var
  ]


def init_mpl():
    """
    It sets the font size of the tick labels to the value of the global variable `ticks_fsize`
    """
    mpl.rcParams['ytick.labelsize'] = ticks_fsize
    mpl.rcParams['xtick.labelsize'] = ticks_fsize

def create_dir(root_dir, dir_path):
    """
    It creates, starting from `root_dir`, that must exist BEFORE, 
    all the subdirectories specified in `dir_path`
    """
    tmp = root_dir
    for sub_dir in dir_path.split('/'):
        tmp += '/' + sub_dir if tmp[-1] != '/' else sub_dir
        makedirs(tmp, exist_ok=True)


def change_Dom_ref(df, newRef = 'MED'):
    """
    Change the Domain reference category of models by putting "0" as prefix.
    Possible values for `newRef` are `'BAL'`, `'MED'`, `'NOR'`
    """
    _df = df.copy()

    
    _df.loc[:, 'Dom'] = _df.Dom.apply(lambda x: '0%s' % x if x == newRef else x)
    return _df

def change_VType_ref(df, newRef = 1):
    """
    Change the VType reference category of models by putting "0" as prefix
    possible values for `newRef` are  `0, 1, ..., 13, 15`
    """
    _newRef = ('_0%d' if newRef < 10 else '_%d') % newRef
    
    _df = df.copy()
    _df.loc[:, 'VType'] = _df.VType.apply(lambda x: '0%s' % x if x == _newRef else x)
    return _df

def cross_validation(data, models_dict, dep_var, nfolds=10, randseed=1, save_dir=''):
    """
    It takes a dataframe, a dictionary of models, and a dependent variable, the number of cv-folds and the random seed,
    then performs a `nfolds`-fold cross validation and saves the results into `save_dir` 
    
    Parameters
    ----------
    data : pandas.DataFrame
        the dataframe containing the data

    models_dict: dictionary
        a dictionary of model formulas, where the key is the model id and the value is the formula 

    dep_var : str
        name of the dependent variable, can be `'Etot'` or `'Eber'`

    nfolds: int (optional)
        number of folds to use in cross validation, defaults to 10 (optional)
    
    randseed: int (optional)
        the random seed, defaults to 1
    
    save_dir: str (optional)
        the directory where you want to save the output
    """
    save = len(save_dir) > 0
    dep_var = models_dict[list(models_dict.keys())[0]].split('~')[0].strip()
    
    scoreDB = pd.DataFrame(
        dict(iteration=[], model_id=[], formula=[], aic=[], bic=[], loglike=[],
             #nfc=[], nrv=[], p=[], 
             R2trM=[], R2trC=[], R2vM=[], R2vC=[],
             RMSEtrM=[], RMSEtrC=[], RMSEvM=[], RMSEvC=[]
        )
      )

    idxs = data.index.to_list()

    # set r and py randseed
    setseed = robjects.r(f"set.seed({randseed})")
    random.seed(randseed)
    
    
    random.shuffle(idxs)

    train_val_dim   = int(len(data))
    train_val_idxs  = idxs

    train_val_idxs_df = pd.DataFrame(dict(idx=train_val_idxs))

    kf = KFold(n_splits=nfolds)
    nit = 1
    for train_idxset, validation_idxset in kf.split(train_val_idxs):

        tidxs = train_val_idxs_df.iloc[train_idxset].values.flatten().tolist()
        vidxs = train_val_idxs_df.iloc[validation_idxset].values.flatten().tolist()

        train_data = data.iloc[tidxs].reset_index(drop=True)
        validation_data = data.iloc[vidxs].reset_index(drop=True)


        print('=====================================================================================')
        print('[%s] Iteration %d/%d started...' % (datetime.datetime.now().time(), nit, nfolds))

        itDB = cross_validation_iteration(
            nit, train_data, validation_data, models_dict
        )
        scoreDB = scoreDB.append(itDB)
        print('[%s] Iteration %d/%d DONE.' % (datetime.datetime.now().time(), nit, nfolds))
        nit += 1


    print('=====================================================================================')
    print('*****************************************************************************************')
    print('                             Cross validation for %s DONE.                         '%dep_var)
    print('*****************************************************************************************')
    # print('\n\n')

    all_iterations_scoresDB = scoreDB.astype(dict(zip(['iteration', 'model_id'], ['int', 'int'])))
    # save all_iterations_scoresDB both in csv
    all_iterations_scoresDB.to_csv(save_dir + 'all_iterations_scoresDB.csv', index=False)

    final_scoresDB = compute_cv_scores_table(all_iterations_scoresDB)
    # save consolidated score DB both in csv and xlsx
    final_scoresDB.to_csv('%sscores_summary.csv'%save_dir, index=False)
    final_scoresDB.to_excel('%sscores_summary.xlsx'%save_dir, index=False)

    
    idxs_best = [final_scoresDB.aic.idxmin(),
                 final_scoresDB.bic.idxmin(),
                 final_scoresDB.loglike.idxmax(),
                 #'--', '--', '--', #nfc, nrv, p columns
                 final_scoresDB.R2trM.idxmax(),
                 final_scoresDB.R2trC.idxmax(),
                 final_scoresDB.R2vM.idxmax(),
                 final_scoresDB.R2vC.idxmax(),
                 final_scoresDB.RMSEtrM.idxmin(),
                 final_scoresDB.RMSEtrC.idxmin(),
                 final_scoresDB.RMSEvM.idxmin(),
                 final_scoresDB.RMSEvC.idxmin(),
                 '--'] #formula column


    vals = final_scoresDB.values

    #print('id\tAIC\tBIC\tloglike\tnfc\tnrv\tp\tR2trM\tR2trC\tR2vM\tR2vC\tRMSEtrM\tRMSEtrC\tRMSEvM\tRMSEvC\tformula')
    # print('id\tAIC\tBIC\tloglike\tR2trM\tR2trC\tR2vM\tR2vC\tRMSEtrM\tRMSEtrC\tRMSEvM\tRMSEvC\tformula')
    
    # for row in vals:
      # rowl = list(row)
      # rowl.append(rowl.pop(1)) #move formula to the end
      
      # scores_str = '\t'.join(map( str, rowl))
      # print(scores_str)

    # print( 'best\t'+'\t'.join(map( str, idxs_best)) )
    
    print('[%s] CV DONE.\nSaving models info...' % datetime.datetime.now().time())
    #save single model output
    save_model_summaries(data, models_dict, dep_var, save_dir)
    print('[%s] Done.\nAll output can be found in %s' % (datetime.datetime.now().time(), save_dir))



def cross_validation_iteration(iteration, train_data_, validation_data_, models_dict):
    """
    This function implements a generic itaration of a k-fold cross-validation procedure.
    It takes a dictionary of models, a training set, and a validation set, and returns a dataframe with
    the model scores
    
    Parameters
    ----------
    
    iteration : int
        iteration number

    train_data_ : pandas.DataFrame
        a dataframe containing the training data
    
    validation_data_ : int
        a dataframe containing the data that will be used for validation

    models_dict : dictionary
        a dictionary of models, where the key is the model ID and the value is the model
    formula


    Returns
    -------

    A dataframe with the scores of the models
    """
    iterations, model_ids, formulae, nfcs, nrvs, ps, aics, bics, llfs, R2TRM, R2TRC, R2VALM, R2VALC, R2TS, RMSETRM, RMSETRC, RMSEVALM, RMSEVALC, RMSETS = [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []

    dep_var = (models_dict[1].split('~')[0]).strip()


    train_data, validation_data = train_data_.copy(), validation_data_.copy()



    for idx, formula in models_dict.items():
      # print('[%s] -------------model %d/%d----------------' % (datetime.datetime.now().time(), idx, len(models_dict)))
      try:
          # print('[%s] Instanciating model %s ...' % (datetime.datetime.now().time(), formula))
          mlm = '|' in formula
          model = Lmer(formula, data=train_data) if mlm else smf.ols(formula, train_data).fit()
          
          # print('[%s] Fitting model...' % datetime.datetime.now().time())
          # print(len(train_data), formula)
          if mlm:
            model.fit(REML=False, summary=False)

          # print('[%s] Evaluating model...' % datetime.datetime.now().time())

          # AIC/BIC https://slideplayer.com/slide/10551973/
          fc = len(model.coefs) if mlm else len(model.params)
          fc -= 1
          rv = len(model.ranef_var) if mlm else 0
          p = fc + rv
          pp1 = p + 1
          ll = model.logLike if mlm else model.llf
          AIC = -2 * ll + 2 * pp1
          BIC = -2 * ll + np.log(len(train_data)) * pp1

          # training-set marginal and conditional predictions
          ytrain_predM = model.predict(train_data, use_rfx=False, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(train_data)
          ytrain_predC = model.predict(train_data, use_rfx=True, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(train_data)
          # validation-set marginal and conditional predictions
          yval_predM = model.predict(validation_data, use_rfx=False, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(validation_data)
          yval_predC = model.predict(validation_data, use_rfx=True, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(validation_data)

          # training-set marginal and conditional R2 and RMSE
          r2trainM = r2_score(train_data[dep_var], ytrain_predM)
          r2trainC = r2_score(train_data[dep_var], ytrain_predC)

          rmseTrainM = mean_squared_error(train_data[dep_var], ytrain_predM, squared=False)
          rmseTrainC = mean_squared_error(train_data[dep_var], ytrain_predC, squared=False)

          # validation-set marginal and conditional R2 and RMSE
          r2valM = r2_score(validation_data[dep_var], yval_predM)
          r2valC = r2_score(validation_data[dep_var], yval_predC)

          rmseValM = mean_squared_error(validation_data[dep_var], yval_predM, squared=False)
          rmseValC = mean_squared_error(validation_data[dep_var], yval_predC, squared=False)


          # append scores to output vectors that will returned as a pd.DataFrame
          iterations.append(iteration)
          model_ids.append(idx)
          formulae.append(formula)
          
          aics.append(AIC)
          bics.append(BIC)
          llfs.append(ll)

          nfcs.append(fc)
          nrvs.append(rv)
          ps.append(p)

          R2TRM.append(r2trainM)
          R2TRC.append(r2trainC)

          R2VALM.append(r2valM)
          R2VALC.append(r2valC)


          RMSETRM.append(rmseTrainM)
          RMSETRC.append(rmseTrainC)

          RMSEVALM.append(rmseValM)
          RMSEVALC.append(rmseValC)

              # print('[%s] Model %d DONE.' % (datetime.datetime.now().time(),idx))

      except Exception as e:
          
          print('[%s] ERROR in model %d, it %d' % (datetime.datetime.now().time(), idx, iteration))
          print(e)

          continue
  

    return pd.DataFrame(dict(
        iteration=iterations, model_id=model_ids, formula=formulae,
        aic=aics, bic=bics, loglike=llfs,
        #nfc=nfcs, nrv=nrvs, p=ps,
        R2trM=R2TRM, R2trC=R2TRC, R2vM=R2VALM, R2vC=R2VALC,
        RMSEtrM=RMSETRM, RMSEtrC=RMSETRC, RMSEvM=RMSEVALM, RMSEvC=RMSEVALC
        
    ))


def compute_cv_scores_table(all_iterations_scoresDB, nsd = 3):

  """
  Computes and returns consolidated cv scores from `all_iteration_scoresDB` dataframe 
  generated by `cross_validation` function.

  Parameters
  ----------
  all_iteration_scoresDB : pandas.DataFrame
      cv scores table generated by `cross_validation` function

  nsd : int, optional
      number of significant digits of scores in final report.
  
  Returns
  -------
  final_scoresDF : pandas.DataFrame
      consolidated cv scores dataframe.

  """
  # one row for each iteration and model
  scoreDB = all_iterations_scoresDB.copy()
  # grouping on model_id and averaging scores
  scores_means = scoreDB.groupby('model_id').mean().drop(columns=['iteration'])
  

  # creating final_scoresDF for reporting table
  final_scoresDF = scores_means

  # define model_id and formula columns that have been lost during grouping
  idxs_models = scores_means.index.to_list()
  formulae = [scoreDB.loc[scoreDB.model_id == mid, 'formula'].unique()[0] for mid in idxs_models]
  final_scoresDF.loc[:, 'model_id'] = idxs_models
  final_scoresDF.loc[:, 'formula'] = formulae

  # transform scores for cleaner reporting
  toint = lambda x: '%d' % x
  r2transform = lambda x: ('%.3f' % x).replace('0.', '.')
  
  #final_scoresDF.nfc = final_scoresDF.nfc.apply(toint).astype(int)
  #final_scoresDF.nrv = final_scoresDF.nrv.apply(toint).astype(int)
  #final_scoresDF.p = final_scoresDF.p.apply(toint).astype(int)
  final_scoresDF.aic = final_scoresDF.aic.round(nsd)
  final_scoresDF.bic = final_scoresDF.bic.round(nsd)
  final_scoresDF.loglike = final_scoresDF.loglike.round(nsd)
  final_scoresDF.R2trM = final_scoresDF.R2trM.round(nsd)
  final_scoresDF.R2trC = final_scoresDF.R2trC.round(nsd)
  final_scoresDF.R2vM = final_scoresDF.R2vM.round(nsd)
  final_scoresDF.R2vC = final_scoresDF.R2vC.round(nsd)
  final_scoresDF.RMSEtrM = final_scoresDF.RMSEtrM.round(nsd)
  final_scoresDF.RMSEtrC = final_scoresDF.RMSEtrC.round(nsd)
  final_scoresDF.RMSEvM = final_scoresDF.RMSEvM.round(nsd)
  final_scoresDF.RMSEvC = final_scoresDF.RMSEvC.round(nsd)

  return final_scoresDF[
    ['model_id', 'formula', 'aic', 'bic', 
     'loglike',
     #'nfc', 'nrv', 'p', 
     'R2trM', 'R2trC', 'R2vM', 'R2vC', 
     'RMSEtrM','RMSEtrC', 'RMSEvM', 'RMSEvC']]

def color_and_border_rows(worksheet, border=None):
    """
    This function takes a worksheet and a border style and colors the rows in the worksheet with
    alternating colors and applies the border style to each cell
    """
    # color group of rows
    #1-5, 21-25 blue
    #6-20, 26-40 orange
    alphabet_upper = string.ascii_uppercase
    cols = alphabet_upper[:alphabet_upper.index('N')]
    for r in range(2, 42):
        for c in cols:
            fgc = 'a8d4eb' if 1 < r < 7 or 21 < r < 27 else 'f6d496'
            worksheet['%s%d' % (c, r)].fill = PatternFill(fgColor=fgc, fill_type="solid")
            if border:
                worksheet['%s%d' % (c, r)].border = border


def set_columns_width(worksheet, wmin=10, wmax=60):
    """
    It sets the width of all the columns but one to `vmax`. The 'B' columns width is set to `vmin`
    """
    # hardcoded col witdh
    # for col, value in dims.items():
    alphabet_upper = string.ascii_uppercase
    cols = alphabet_upper[:alphabet_upper.index('N')]
    for col in cols:
        worksheet.column_dimensions[col].width = wmax if col == 'B' else wmin
        # ws.column_dimensions[col].alignment = Alignment(horizontal='center', vertical='center')


def center_align_excel(work_sheet):
    """
    It takes a worksheet as an argument and then iterates through 
    each cell setting the alignment to center.
    """
    for col in work_sheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def save_model_summaries(data, models_dict, dep_var, outdir):
  """
  Save cv scores in final excel with a summary sheet and one sheet per model.
  Hyperlinks are available to facilitate the navigation.
  """
    
  xlsx_path_in  = '%sscores_summary.xlsx' % outdir
  xlsx_path_out = '%s%s.xlsx' % (outdir, dep_var)


  final_scoresDF = pd.read_excel(xlsx_path_in, engine='openpyxl')

  writer = pd.ExcelWriter(xlsx_path_out, engine='openpyxl')

  final_scoresDF.to_excel(writer, index=False, sheet_name='CVScores')
  writer.save()

  wb = writer.book
  cvsb = wb['CVScores']
  idx_offset = 1

  # freeze model id, formula cols and header row
  cvsb.freeze_panes = "C2"

  thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
  # fix layout
  color_and_border_rows(cvsb, border=thin_border)
  set_columns_width(cvsb)
  center_align_excel(cvsb)

  for idx, formula in models_dict.items():
    try:
      sidx = ('mod0%d' if idx < 10 else 'mod%d') % idx
      fit_evaluate_single_model(idx, formula, data, xlsx_writer=writer)
      
      # cvscores -> model link
      ws = wb['CVScores']
      cellid = ws.cell(row=idx + idx_offset, column=1)
      # cellf = ws.cell(row=idx + idx_offset, column=2)
      cellid.hyperlink = "#%s!A1" % (sidx)

      cp = copy.copy(cellid.font)
      cp.underline = 'single'
      cp.color     = 'FF0000FF'

      cellid.font    = cp

      # model -> cvscores link
      ws = wb[sidx]
      cell = ws.cell(row=1, column=1)

      cell.hyperlink = '#CVScores!A1'
      cell.value = 'to CVScores'

      del cp
      #font properties
      cp = copy.copy(cell.font)
      cp.italic    = 'True'
      cp.bold      = 'True'
      cp.underline = 'single'
      cp.color     = 'FF0000FF'

      cell.font    = cp


      # center cell content (but for fromula cell) and wider first column
      center_align_excel(ws)
      ws.cell(row=3, column=2).alignment = Alignment(horizontal='left')
      ws.column_dimensions['A'].width = 25

    except Exception as e:
      print('ERROR in model %d' % idx)
      print(e)
      idx_offset -= 1
      continue
    
  writer.save()
  writer.close()


def fit_evaluate_single_model(model_id, formula, data_to_fit, xlsx_writer=None):
    """
    It fits a model, evaluates it, and writes the results to an Excel file
    """

    if not xlsx_writer:
        print('no output file writer provided')
        return

    dep_var = formula.split('~')[0].strip()
    isLn = dep_var[:2] == 'ln'
    mlm = '|' in formula

    data = data_to_fit

    model = Lmer(formula, data=data) if mlm else smf.ols(formula, data).fit()

    if mlm:
        model.fit(summary=False, REML=False)

    yobs = data[dep_var]
    ypredM = model.predict(data, use_rfx=False, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(data)
    ypredC = model.predict(data, use_rfx=True, verify_predictions=False, skip_data_checks=True) if mlm else model.predict(data)

    fc = len(model.coefs) if mlm else len(model.params)
    fc -= 1
    rv = len(model.ranef_var) if mlm else 0
    p = fc + rv
    pp1 = p + 1
    ll = model.logLike if mlm else model.llf
    AIC = -2 * ll + 2 * pp1
    BIC = -2 * ll + np.log(len(data)) * pp1

    r2M = r2_score(yobs, ypredM)
    r2C = r2_score(yobs, ypredC)
    rmseM = mean_squared_error(yobs, ypredM, squared=False)
    rmseC = mean_squared_error(yobs, ypredC, squared=False)

    sheet_name = ('mod0%d' if model_id < 10 else 'mod%d') % model_id
    descr_index = ['model_id', 'formula', 'nobs']

    tmp_gk = list(model.grps.keys()) if mlm else []  # keys in gropus dict
    tmp_g = tmp_gk[0] if mlm else ''

    descr_data = [model_id, formula, len(data)]

    if mlm:
        descr_index.append('ngroups')
        descr_data.append('%d' % model.grps[tmp_g])

    descr_df = pd.DataFrame(index=descr_index, columns=[''], data=descr_data)
    descr_df.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=0)

    nrow = len(descr_index) + 2

    pd.DataFrame(index=['in-sample scores'], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
    nrow += 2

    scores_df = pd.DataFrame(
        columns=['', 'AIC', 'BIC', 'LogLike',
                 #'nfc', 'nrv', 'p',
                 'R2m', 'R2c', 'RMSEm', 'RMSEc'],
        data=[[np.nan, AIC, BIC, ll,
              #fc, rv, p,
              r2M, r2C, rmseM, rmseC]]
    )

    scores_df.AIC = scores_df.AIC.astype(int)
    scores_df.BIC = scores_df.BIC.astype(int)
    scores_df.LogLike = scores_df.LogLike.astype(int)
    #scores_df.nfc = scores_df.nfc.astype(int)
    #scores_df.nrv = scores_df.nrv.astype(int)
    #scores_df.p = scores_df.p.astype(int)
    scores_df.RMSEm = scores_df.RMSEm.astype(int)
    scores_df.RMSEc = scores_df.RMSEc.astype(int)
    scores_df.R2m = scores_df.R2m.round(3)
    scores_df.R2c = scores_df.R2c.round(3)

    scores_df.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow, index=False)

    nrow += 4  # 2(scores) + 2(empty)

    if mlm:
        fcs = model.coefs.drop(columns=['DF', 'Sig'])

        fcs.Estimate = fcs.Estimate.round(1)
        fcs['2.5_ci'] = fcs['2.5_ci'].round(1)
        fcs['97.5_ci'] = fcs['97.5_ci'].round(1)
        fcs.SE = fcs.SE.round(1)
        fcs['T-stat'] = fcs['T-stat'].round(2)
        fcs['P-val'] = fcs['P-val'].round(3)

        pd.DataFrame(index=['Fixed effects'], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += 2
        fcs.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += len(fcs) + 2

        randcfs      = model.ranef.round(1)
        randcfs_vars = model.ranef_var.round(1)

        pd.DataFrame(index=['Random effects variances'], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += 2
        randcfs_vars.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += len(randcfs_vars) + 2

        # pd.DataFrame(index=['Random effects'], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        # nrow += 2
        # randcfs.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        # nrow += len(randcfs) + 2
    else:
        fe_data = [[pr, ci, ci1, se, tv, pv]
                   for pr, ci, ci1, se, tv, pv in zip(
                model.params.round(1).values, model.conf_int().round(1).values[:, 0],
                model.conf_int().round(1).values[:, 1], model.bse.round(1).values,
                model.tvalues.round(2).values, model.pvalues.round(3).values)]

        fe_df = pd.DataFrame(
            index=model.params.index,
            columns=['Estimate', '2.5_ci', '97.5_ci', 'SE', 'T-stat', 'P-val'],
            data=fe_data
        )

        pd.DataFrame(index=['Fixed effects'], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += 2
        fe_df.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow)
        nrow += len(fe_df) + 2

    # nrow += 2
    # pd.DataFrame(index=['predicted %s' % dep_var], columns=[''], data=['']).to_excel(xlsx_writer, sheet_name=sheet_name,
                                                                            # startrow=nrow)
    # nrow += 2
    # ypredDF = pd.DataFrame(dict(IMON=data.IMOn, marg=ypredM, cond=ypredC))
    # ypredDF.loc[:, dep_var] = data[dep_var]
    # ypredDF.to_excel(xlsx_writer, sheet_name=sheet_name, startrow=nrow, index=True)

    xlsx_writer.save()

    return model


def reformat_ticks(ax, xminor_multiple=None, xmajor_multiple=None):
    """
    It removes the right and top spines, sets the major and minor ticks, and sets the major tick labels
    """
    ax.spines['right'].set_visible(False)
    # ax.spines['left'].set_visible(False)
    # ax.spines['bottom'].set_visible(False)
    ax.spines['top'].set_visible(False)

    if xminor_multiple and xmajor_multiple:
        ax.xaxis.set_major_locator(MultipleLocator(xmajor_multiple))
        def xfmt(val, pos):
            return '%d' % val
        ax.xaxis.set_major_formatter(FuncFormatter(xfmt))
        # For the minor ticks, use no labels; default NullFormatter.
        ax.xaxis.set_minor_locator(MultipleLocator(xminor_multiple))

    ax.tick_params(which='both', width=1.2)
    ax.tick_params(which='major', length=7)
    ax.tick_params(which='minor', length=3.5, color='gray')


def cv_results_plot(variable, chosen_model=30, cv_dir='', outdir='', save_plot=False):
    """
    It plots the AIC and RMSE values for each model in the cross-validation as in [1].

    References
    ----------

    [1] Mannarini G, Salinas ML, Carelli L, Fassò A. 
    How COVID-19 Affected GHG Emissions of Ferries in Europe. 
    Sustainability. 2022; 14(9):5287. https://doi.org/10.3390/su14095287 
    """
    def format_func_e4(value, tick_number):
        v = np.round(value / 1e4, 2)
        return v

    def format_func_e3(value, tick_number):
        den = 1e3 if variable == 'Etot' else 1e2
        rnd = 0 if variable == 'Etot' else 1
        v = np.round(value / den, rnd)
        return int(v) if variable == 'Etot' else v

    init_mpl()

    fsize_incr = 7
    msize_incr = 150

    scores_dir = '%s%s/' % (cv_dir, variable)
    data = pd.read_csv('%sscores_summary.csv' % scores_dir )

    fig, ax = plt.subplots(figsize=figsize)
    ax.grid(axis='both', zorder=0, color='lightgray')
    ax1 = ax.twinx()

    title = '%s' % variable
    ax.set_title(title, fontsize=title_fsize*1.1, pad=25)
    
    
    aic = data.aic
    logl = data.loglike
    rmse = data.RMSEvC
    nmodels = len(aic)

    x = list(range(1, nmodels+1))


    # mask bests
    aicMin = aic.min()
    aMinIdx = aic.idxmin()
    aic[aMinIdx] = np.nan

    # aic plot
    ax.scatter(x, aic, zorder=10, s=40 + msize_incr * .25, edgecolor='black', facecolor='None', label='AIC')

    # plot best aic here
    ax.scatter(aMinIdx + 1, aicMin, zorder=10, s=40 + msize_incr * .25, edgecolor='black', facecolor='black')  # ,
    # label='min AIC')

    xlm = ax.get_xlim()
    ylm = ax.get_ylim()

    ax.set_ylabel('AIC [$\mathrm{10^4}$]', fontsize=label_fsize + fsize_incr)
    ax1.set_ylabel('RMSE [kt]', fontsize=label_fsize + fsize_incr)
    ax.set_xlabel('model id', fontsize=label_fsize + fsize_incr)


    ylim = (14800, 15800) if variable == 'Eber' else (18000, 20100)
    ylim1 = (1150, 1950) if variable == 'Eber' else (8000, 27000) # median VTypes

    props = {'ha': 'center', 'va': 'bottom'}
    ax.vlines(chosen_model, ylim[0], ylim[1] * 10, color='black', zorder=5, linestyles='--')

    bestRmseId = rmse.idxmin() + 1
    bestRmse = rmse[rmse.idxmin()]
    rmse[bestRmseId - 1] = np.nan

    # plot rmse on sharedy axis
    ax1.scatter(x, rmse, zorder=10, s=70, marker='d', edgecolor='black', facecolor='None', label='RMSEvC')
    # for legend
    ax.scatter(-np.array(x), rmse, zorder=10, s=70, marker='d', edgecolor='black', facecolor='None', label='RMSEvC')

    # best rmsevm
    ax1.scatter(bestRmseId, bestRmse, zorder=10, s=70, edgecolor='black', facecolor='black',
                marker='d', label='min RMSEvC')

    ax.set_xlim(xlm)
    ax.set_ylim(ylim)
    ax1.set_ylim(ylim1)


    reformat_ticks(ax, xmajor_multiple=5, xminor_multiple=1)
    # else:
    # ax.ticklabel_format(style='sci', axis='y', scilimits=(0, 0))
    ax.yaxis.offsetText.set_fontsize(ticks_fsize + fsize_incr)
    # Create legend & Show graphic
    handles, labels = ax.get_legend_handles_labels()
    # handles = handles[1:] + [handles[0]]
    # labels = labels[1:] + [labels[0]]

    legend = ax.legend(handles, labels, fontsize=txt_fsize * .6 + fsize_incr, ncol=1,
                       facecolor="white",
                       loc='upper right')

    legend.get_frame().set_alpha(1)
    legend.get_frame().set_facecolor('white')

    for xmaj in ax.xaxis.get_majorticklocs():
        ax.axvline(x=xmaj, ls='--', color='darkgray', linewidth=.7)
    for ymaj in ax.yaxis.get_majorticklocs():
        ax.axhline(y=ymaj, ls='--', color='darkgray', linewidth=.7)
    for ymin in ax.yaxis.get_minorticklocs():
        ax.axhline(y=ymin, ls='--', color='darkgray', linewidth=.5)

    ax.yaxis.set_major_formatter(plt.FuncFormatter(format_func_e4))
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(format_func_e3))

    plt.tight_layout()

    if save_plot:
        vr = variable
        plt.savefig('%s%s_cv_results.pdf' % (outdir, vr))
        plt.savefig('%s%s_cv_results.png' % (outdir, vr))
        print('%s cv results plot saved in png/pdf:\n%s_cv_results<.fmt>' % (variable, outdir+variable))

    plt.show()
